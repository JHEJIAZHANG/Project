import csv
import io
from typing import List, Dict


def parse_courses_csv(data: bytes, encoding: str = "utf-8") -> List[Dict]:
    """Parse CSV bytes to course dicts.

    Expected headers (case-insensitive, optional):
    - title (required)
    - description
    - instructor
    - classroom
    """
    text = data.decode(encoding, errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    out: List[Dict] = []
    for row in reader:
        if not row:
            continue
        title = (row.get("title") or row.get("Title") or row.get("課程名稱") or "").strip()
        if not title:
            continue
        out.append({
            "title": title,
            "description": (row.get("description") or row.get("Description") or row.get("說明") or "").strip(),
            "instructor": (row.get("instructor") or row.get("Instructor") or row.get("教師") or "").strip(),
            "classroom": (row.get("classroom") or row.get("Classroom") or row.get("教室") or "").strip(),
        })
    return out


def parse_courses_ical(data: bytes) -> List[Dict]:
    """Parse iCalendar bytes to course dicts using VEVENT summary/description/location.
    We only extract basic metadata for course creation.
    """
    from icalendar import Calendar

    cal = Calendar.from_ical(data)
    out: List[Dict] = []
    for component in cal.walk():
        if component.name != "VEVENT":
            continue
        title = str(component.get("SUMMARY", "")).strip()
        if not title:
            continue
        desc = str(component.get("DESCRIPTION", "")).strip()
        loc = str(component.get("LOCATION", "")).strip()
        out.append({
            "title": title,
            "description": desc,
            "instructor": "",
            "classroom": loc,
        })
    return out


def parse_courses_xlsx(data: bytes) -> List[Dict]:
    """Parse Excel (.xlsx) bytes to course dicts.
    支援表頭（不分大小寫/中英）：title/課程名稱、description/說明、instructor/教師、classroom/教室。
    若讀不到表頭，會嘗試使用首行作為表頭。
    若完全無法解析且已設定 Gemini_API_KEY，回退以 Gemini 協助抽取結構。
    """
    out: List[Dict] = []
    try:
        from openpyxl import load_workbook  # type: ignore
        import io as _io
        wb = load_workbook(filename=_io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return out
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        # 若表頭全空，視為無表頭，嘗試以固定欄位序：title, description, instructor, classroom
        has_header = any(h for h in headers)
        if has_header:
            mapping = {}
            for idx, h in enumerate(headers):
                key = (h or "").strip().lower()
                if key in {"title", "課程名稱", "name"}:
                    mapping["title"] = idx
                elif key in {"description", "說明", "desc"}:
                    mapping["description"] = idx
                elif key in {"instructor", "教師", "teacher"}:
                    mapping["instructor"] = idx
                elif key in {"classroom", "教室", "location"}:
                    mapping["classroom"] = idx
            data_rows = rows[1:]
            for r in data_rows:
                title = str(r[mapping.get("title", -1)]).strip() if mapping.get("title", -1) >= 0 and r[mapping.get("title", -1)] is not None else ""
                if not title:
                    continue
                desc = str(r[mapping.get("description", -1)]).strip() if mapping.get("description", -1) >= 0 and r[mapping.get("description", -1)] is not None else ""
                instr = str(r[mapping.get("instructor", -1)]).strip() if mapping.get("instructor", -1) >= 0 and r[mapping.get("instructor", -1)] is not None else ""
                room = str(r[mapping.get("classroom", -1)]).strip() if mapping.get("classroom", -1) >= 0 and r[mapping.get("classroom", -1)] is not None else ""
                out.append({
                    "title": title[:255],
                    "description": desc[:1000],
                    "instructor": instr[:100],
                    "classroom": room[:100],
                })
        else:
            for r in rows:
                vals = [(str(v).strip() if v is not None else "") for v in r]
                if not any(vals):
                    continue
                title = vals[0]
                if not title:
                    continue
                desc = vals[1] if len(vals) > 1 else ""
                instr = vals[2] if len(vals) > 2 else ""
                room = vals[3] if len(vals) > 3 else ""
                out.append({
                    "title": title[:255],
                    "description": desc[:1000],
                    "instructor": instr[:100],
                    "classroom": room[:100],
                })
        if out:
            return out
    except Exception as e:
        # 記錄但不拋出，讓後備流程接手
        print(f"[xlsx] parse error: {e}")

    # 後備：Gemini 協助從表格文字抽取（將每列串為文字）
    try:
        import os, json
        if os.getenv("Gemini_API_KEY") or os.getenv("GEMINI_API_KEY"):
            from services.gemini_client import get_model  # type: ignore
            model = get_model("gemini-1.5-flash")
            system = (
                "Extract course list from the given table content. Output STRICT JSON only: "
                "{\"items\":[{\"title\":string,\"description\":string,\"instructor\":string,\"classroom\":string}]}"
            )
            # 為避免複雜處理，直接將二維表格轉純文字給 Gemini
            lines = []
            try:
                from openpyxl import load_workbook  # type: ignore
                import io as _io
                wb = load_workbook(filename=_io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    vals = [(str(v).strip() if v is not None else "") for v in row]
                    if any(vals):
                        lines.append("\t".join(vals))
            except Exception as e:
                print(f"[xlsx] load workbook for gemini text failed: {e}")
            table_text = "\n".join(lines)
            resp = model.generate_content([
                {"role": "user", "parts": [system, table_text]}
            ])
            content = (resp.text or "").strip()
            if content.startswith("```json") and content.endswith("```"):
                content = content[7:-3].strip()
            data = json.loads(content)
            for it in data.get("items", []) or []:
                title = (it.get("title") or "").strip()
                if not title:
                    continue
                out.append({
                    "title": title[:255],
                    "description": (it.get("description") or "")[:1000],
                    "instructor": (it.get("instructor") or "")[:100],
                    "classroom": (it.get("classroom") or "")[:100],
                })
    except Exception as e:
        print(f"[xlsx] gemini fallback error: {e}")

    return out


def parse_courses_xls(data: bytes) -> List[Dict]:
    """Parse legacy Excel (.xls) bytes to course dicts using xlrd.
    欄位對應與 .xlsx 相同。
    """
    out: List[Dict] = []
    print(f"[xls] 開始解析，檔案大小: {len(data)} bytes")
    
    # 偵測偽 Excel（其實是 HTML 表格）
    try:
        sniff = data[:200].lower()
        if b"<html" in sniff or b"<table" in sniff or b"<td" in sniff:
            print("[xls] 偵測到 HTML 內容，使用 HTML 解析器")
            # 將 HTML 轉純文字，再重用課表文字解析
            try:
                html = data.decode("utf-8", errors="ignore")
            except Exception:
                html = data.decode("big5", errors="ignore")
            import re
            # 保留換行
            html = re.sub(r"<br\s*/?>", "\n", html, flags=re.I)
            # 將單元格關閉視為換行，避免內容黏在一起
            html = re.sub(r"</td>", "\n", html, flags=re.I)
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " \n", text)
            try:
                items = parse_multiple_courses_from_timetable(text)
                print(f"[xls] HTML 解析到 {len(items)} 個項目")
            except Exception as e:
                print(f"[xls] HTML 解析失敗: {e}")
                items = []
            # 僅回傳基本欄位（title/instructor/classroom/description），時間表先不寫入
            basic: List[Dict] = []
            for it in items:
                title = (it.get("title") or it.get("name") or "").strip()
                if not title:
                    continue
                basic.append({
                    "title": title[:255],
                    "description": (it.get("description") or "")[:1000],
                    "instructor": (it.get("instructor") or "")[:100],
                    "classroom": (it.get("classroom") or "")[:100],
                })
            if basic:
                print(f"[xls] HTML 解析成功，回傳 {len(basic)} 個課程")
                return basic
    except Exception as e:
        print(f"[xls] html sniff error: {e}")

    try:
        import xlrd  # type: ignore
        print("[xls] 使用 xlrd 解析器")
        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        print(f"[xls] 工作表行數: {sheet.nrows}, 列數: {sheet.ncols}")
        if sheet.nrows == 0:
            return out
        headers = [(str(sheet.cell_value(0, c)).strip() if sheet.cell_value(0, c) is not None else "") for c in range(sheet.ncols)]
        print(f"[xls] 標題行: {headers}")
        has_header = any(headers)
        if has_header:
            mapping = {}
            for idx, h in enumerate(headers):
                key = (h or "").strip().lower()
                if key in {"title", "課程名稱", "name"}:
                    mapping["title"] = idx
                elif key in {"description", "說明", "desc"}:
                    mapping["description"] = idx
                elif key in {"instructor", "教師", "teacher"}:
                    mapping["instructor"] = idx
                elif key in {"classroom", "教室", "location"}:
                    mapping["classroom"] = idx
            print(f"[xls] 欄位對應: {mapping}")
            for r in range(1, sheet.nrows):
                rowvals = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                title = str(rowvals[mapping.get("title", -1)]).strip() if mapping.get("title", -1) >= 0 else ""
                if not title:
                    continue
                desc = str(rowvals[mapping.get("description", -1)]).strip() if mapping.get("description", -1) >= 0 else ""
                instr = str(rowvals[mapping.get("instructor", -1)]).strip() if mapping.get("instructor", -1) >= 0 else ""
                room = str(rowvals[mapping.get("classroom", -1)]).strip() if mapping.get("classroom", -1) >= 0 else ""
                out.append({
                    "title": title[:255],
                    "description": desc[:1000],
                    "instructor": instr[:100],
                    "classroom": room[:100],
                })
        else:
            print("[xls] 無標題行，使用預設欄位順序")
            for r in range(sheet.nrows):
                rowvals = [(str(sheet.cell_value(r, c)).strip() if sheet.cell_value(r, c) is not None else "") for c in range(sheet.ncols)]
                if not any(rowvals):
                    continue
                title = rowvals[0]
                if not title:
                    continue
                desc = rowvals[1] if len(rowvals) > 1 else ""
                instr = rowvals[2] if len(rowvals) > 2 else ""
                room = rowvals[3] if len(rowvals) > 3 else ""
                out.append({
                    "title": title[:255],
                    "description": desc[:1000],
                    "instructor": instr[:100],
                    "classroom": room[:100],
                })
        print(f"[xls] xlrd 解析成功，回傳 {len(out)} 個課程")
    except Exception as e:
        print(f"[xls] parse error: {e}")
        import traceback
        traceback.print_exc()
    return out
def parse_courses_xml(data: bytes) -> List[Dict]:
    """Parse XML bytes to course dicts.
    支援常見欄位：title/name、description/desc、instructor/teacher、classroom/location。
    若無法解析任何課程且已設定 Gemini_API_KEY，嘗試以 Gemini 進行結構化抽取。
    """
    import xml.etree.ElementTree as ET

    text = data.decode("utf-8", errors="ignore")
    out: List[Dict] = []

    try:
        root = ET.fromstring(text)
    except Exception:
        root = None

    def _pick(elem, names):
        # 先找子元素，再找屬性
        for n in names:
            node = elem.find(n)
            if node is not None and (node.text or "").strip():
                return (node.text or "").strip()
        for n in names:
            val = elem.attrib.get(n)
            if val:
                return val.strip()
        return ""

    if root is not None:
        # 常見集合節點：course/item/entry/row/課程
        candidates = []
        for tag in ["course", "item", "entry", "row", "Course", "Item", "課程"]:
            candidates.extend(root.findall(f".//{tag}"))
        if not candidates:
            candidates = list(root)

        for elem in candidates:
            title = _pick(elem, ["title", "name", "Title", "Name", "課程名稱"]) or ""
            desc = _pick(elem, ["description", "desc", "Description", "說明"]) or ""
            instructor = _pick(elem, ["instructor", "teacher", "Instructor", "Teacher", "教師"]) or ""
            classroom = _pick(elem, ["classroom", "location", "Location", "教室"]) or ""
            if not title:
                # 使用最長文字作為候選標題
                try:
                    texts = [t.strip() for t in elem.itertext() if (t or "").strip()]
                    title = max(texts, key=len, default="").strip()
                except Exception:
                    title = ""
            if title:
                out.append({
                    "title": title[:255],
                    "description": desc[:1000],
                    "instructor": instructor[:100],
                    "classroom": classroom[:100],
                })

    if out:
        return out

    # 後備：Gemini 結構化抽取
    try:
        import os, json
        if os.getenv("Gemini_API_KEY") or os.getenv("GEMINI_API_KEY"):
            from services.gemini_client import get_model  # type: ignore
            model = get_model("gemini-1.5-flash")
            system = (
                "Extract course list from the given XML. Output STRICT JSON only: "
                "{\"items\":[{\"title\":string,\"description\":string,\"instructor\":string,\"classroom\":string}]}"
            )
            resp = model.generate_content([
                {"role": "user", "parts": [system, text]}
            ])
            content = (resp.text or "").strip()
            if content.startswith("```json") and content.endswith("```"):
                content = content[7:-3].strip()
            data = json.loads(content)
            for it in data.get("items", []) or []:
                title = (it.get("title") or "").strip()
                if not title:
                    continue
                out.append({
                    "title": title[:255],
                    "description": (it.get("description") or "")[:1000],
                    "instructor": (it.get("instructor") or "")[:100],
                    "classroom": (it.get("classroom") or "")[:100],
                })
    except Exception as e:
        print(f"[xml] gemini fallback error: {e}")

    return out
def parse_course_from_text(text: str) -> Dict:
    """Parse a free-form OCR text into a course dict (best-effort).
    Heuristics:
    - Prefer labeled lines like: 課程名稱/Title, 教師/Instructor, 教室/Location/Classroom, 說明/Description
    - Fallback: first non-empty line as title; the rest become description.
    """
    import re

    lines = [ln.strip() for ln in (text or "").splitlines()]
    lines = [ln for ln in lines if ln]

    title = ""
    instructor = ""
    classroom = ""
    desc_parts: List[str] = []

    patterns = [
        ("title", re.compile(r"^(課程名稱|課程|title|course)\s*[:：]\s*(.+)$", re.I)),
        ("instructor", re.compile(r"^(教師|老師|instructor|teacher)\s*[:：]\s*(.+)$", re.I)),
        ("classroom", re.compile(r"^(教室|地點|location|classroom)\s*[:：]\s*(.+)$", re.I)),
        ("description", re.compile(r"^(說明|描述|description)\s*[:：]\s*(.+)$", re.I)),
    ]

    consumed = set()
    for idx, ln in enumerate(lines):
        for key, pat in patterns:
            m = pat.match(ln)
            if m:
                val = m.group(2).strip()
                if key == "title" and not title:
                    title = val
                    consumed.add(idx)
                elif key == "instructor" and not instructor:
                    instructor = val
                    consumed.add(idx)
                elif key == "classroom" and not classroom:
                    classroom = val
                    consumed.add(idx)
                elif key == "description":
                    desc_parts.append(val)
                    consumed.add(idx)
                break

    # Fallbacks
    if not title and lines:
        title = lines[0]
        consumed.add(0)
    # Remaining lines become description
    for idx, ln in enumerate(lines):
        if idx not in consumed:
            desc_parts.append(ln)

    description = "\n".join(desc_parts).strip()
    return {
        "title": title[:255],
        "description": description,
        "instructor": instructor[:100],
        "classroom": classroom[:100],
    }


def parse_multiple_courses_from_timetable(text: str) -> List[Dict]:
    """Enhanced parser for weekly timetable OCR text to multiple course dicts.
    
    Handles both horizontal and vertical timetable formats.
    Strategy:
    1. Detect time patterns and course information
    2. Parse day-of-week markers
    3. Extract course names, times, and locations
    4. Group by course and create schedule entries
    """
    import re
    from datetime import datetime
    
    txt = (text or '').replace('\r', '\n').strip()
    if not txt:
        return []
    
    courses = []
    lines = [ln.strip() for ln in txt.split('\n') if ln.strip()]
    
    # Enhanced patterns for Chinese timetables
    day_patterns = {
        '週一': 0, '星期一': 0, '一': 0, 'Monday': 0,
        '週二': 1, '星期二': 1, '二': 1, 'Tuesday': 1,
        '週三': 2, '星期三': 2, '三': 2, 'Wednesday': 2,
        '週四': 3, '星期四': 3, '四': 3, 'Thursday': 3,
        '週五': 4, '星期五': 4, '五': 4, 'Friday': 4,
        '週六': 5, '星期六': 5, '六': 5, 'Saturday': 5,
        '週日': 6, '星期日': 6, '日': 6, 'Sunday': 6,
    }
    
    # Time patterns (HH:MM-HH:MM, HH:MM HH:MM, etc.)
    time_pattern = re.compile(r'(\d{1,2}:\d{2})\s*[-~—至到]\s*(\d{1,2}:\d{2})')
    single_time_pattern = re.compile(r'(\d{1,2}:\d{2})')
    
    # Course name patterns (Chinese characters + alphanumeric)
    course_name_pattern = re.compile(r'[\u4e00-\u9fa5][\u4e00-\u9fa5A-Za-z0-9\s\-\(\)]{2,30}')
    
    # Classroom patterns
    classroom_patterns = [
        re.compile(r'(電\d{3}[^(]*\([^)]+\))', re.I),  # 電708(承曦樓)
        re.compile(r'(電\d{3})', re.I),  # 電101, 電203
        re.compile(r'(\d{3}教室)', re.I),  # 101教室
        re.compile(r'(教\d{3})', re.I),  # 教101
        re.compile(r'(實驗室\d*)', re.I),  # 實驗室, 實驗室1
        re.compile(r'(\d{3})', re.I),  # 101, 203
        re.compile(r'(教室)', re.I),  # 教室
    ]
    
    # Instructor patterns
    instructor_pattern = re.compile(r'([^\s]{2,4}[P]?)\s*$')  # 蔡文隆P, 許晉龍
    
    # Check if this is a vertical timetable format (like the test image)
    has_vertical_format = any('節次' in line or '第一節' in line for line in lines)
    
    if has_vertical_format:
        return _parse_vertical_timetable(lines, day_patterns, time_pattern, 
                                       course_name_pattern, classroom_patterns, 
                                       instructor_pattern)
    else:
        return _parse_horizontal_timetable(lines, day_patterns, time_pattern, 
                                         course_name_pattern, classroom_patterns, 
                                         instructor_pattern)


def _parse_vertical_timetable(lines, day_patterns, time_pattern, course_name_pattern, 
                            classroom_patterns, instructor_pattern):
    """Parse vertical timetable format where courses are listed in columns."""
    import re
    
    courses = []
    current_courses = {}  # title -> course_data
    
    # Enhanced course name pattern for better Chinese recognition
    enhanced_course_pattern = re.compile(r'[\u4e00-\u9fa5]{2,20}(?:應用|設計|實務|專題|安全|倫理|法律|管理|智慧|人工|多媒體|電腦|資訊)')
    
    # Enhanced instructor pattern
    enhanced_instructor_pattern = re.compile(r'([\u4e00-\u9fa5]{2,4}[P]?)\s*$')
    
    # Enhanced classroom pattern
    enhanced_classroom_pattern = re.compile(r'(電\d{3}[^(]*\([^)]+\)|藝\d{3}|教\d{3}|實驗室\d*|\d{3}教室)')
    
    # Find day markers and course information
    current_day = -1
    course_blocks = []
    
    # Process lines to find course blocks
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Check for day markers
        for day_str, day_num in day_patterns.items():
            if day_str in line:
                current_day = day_num
                break
        
        # Look for course information blocks
        if current_day != -1 and line:
            # Check if this line contains course information
            if (enhanced_course_pattern.search(line) or 
                enhanced_instructor_pattern.search(line) or 
                enhanced_classroom_pattern.search(line)):
                
                # Collect course information in this block
                course_info = []
                j = i
                while (j < len(lines) and 
                       lines[j].strip() and 
                       not any(day_str in lines[j] for day_str in day_patterns.keys()) and
                       not lines[j].strip().startswith('第') and  # Skip period markers
                       not re.match(r'^\d{1,2}:\d{2}$', lines[j].strip())):  # Skip time markers
                    if lines[j].strip():
                        course_info.append(lines[j].strip())
                    j += 1
                
                if course_info:
                    course_blocks.append((current_day, course_info))
                i = j - 1  # Skip processed lines
        
        i += 1
    
    # Process course blocks with improved logic
    for day, course_info in course_blocks:
        # Extract course name, instructor, and classroom
        course_name = ""
        instructor = ""
        classroom = ""
        
        # Process each piece of information
        for info in course_info:
            info = info.strip()
            if not info:
                continue
                
            # Check for course name (usually longer Chinese text)
            if not course_name and enhanced_course_pattern.search(info):
                course_name = enhanced_course_pattern.search(info).group(0).strip()
            
            # Check for instructor (usually 2-4 Chinese characters with optional P)
            elif not instructor and enhanced_instructor_pattern.search(info):
                instructor = enhanced_instructor_pattern.search(info).group(1).strip()
            
            # Check for classroom (usually contains numbers and specific keywords)
            elif not classroom and enhanced_classroom_pattern.search(info):
                classroom = enhanced_classroom_pattern.search(info).group(1).strip()
        
        # Only create course if we have a valid course name
        if course_name and len(course_name) >= 2:
            # Clean up course name (remove extra characters)
            course_name = re.sub(r'[^\u4e00-\u9fa5A-Za-z0-9\s\-\(\)]', '', course_name).strip()
            
            # Clean up instructor
            if instructor:
                instructor = re.sub(r'[^\u4e00-\u9fa5A-Za-z]', '', instructor).strip()
            
            # Clean up classroom
            if classroom:
                classroom = re.sub(r'[^\u4e00-\u9fa5A-Za-z0-9\(\)]', '', classroom).strip()
            
            # Create or update course
            if course_name not in current_courses:
                current_courses[course_name] = {
                    'title': course_name,
                    'description': '',
                    'instructor': instructor,
                    'classroom': classroom,
                    'schedule': []
                }
            
            # Add schedule entry
            schedule_entry = {
                'dayOfWeek': day,
                'startTime': '08:10',  # Default start time
                'endTime': '17:10'     # Default end time
            }
            current_courses[course_name]['schedule'].append(schedule_entry)
    
    # Convert to list and filter out invalid courses
    for course_data in current_courses.values():
        if (course_data['schedule'] and 
            course_data['title'] and 
            len(course_data['title']) >= 2):
            courses.append({
                'title': course_data['title'][:255],
                'description': course_data['description'][:1000],
                'instructor': course_data['instructor'][:100],
                'classroom': course_data['classroom'][:100],
                'schedule': course_data['schedule']
            })
    
    return courses[:15]


def _parse_horizontal_timetable(lines, day_patterns, time_pattern, course_name_pattern, 
                              classroom_patterns, instructor_pattern):
    """Parse horizontal timetable format (original logic)."""
    courses = []
    current_courses = {}  # title -> course_data
    current_day = -1
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Check for day markers
        for day_str, day_num in day_patterns.items():
            if day_str in line:
                current_day = day_num
                line = re.sub(rf'^{re.escape(day_str)}\s*', '', line).strip()
                break
        
        if current_day == -1:
            continue
        
        # Look for time patterns
        time_match = time_pattern.search(line)
        if time_match:
            start_time, end_time = time_match.groups()
            course_text = line[time_match.end():].strip()
            
            course_name = ""
            course_match = course_name_pattern.search(course_text)
            if course_match:
                course_name = course_match.group(0).strip()
            
            if course_name:
                classroom = ""
                for pattern in classroom_patterns:
                    classroom_match = pattern.search(course_text)
                    if classroom_match:
                        classroom = classroom_match.group(0)
                        break
                
                instructor = ""
                instructor_match = instructor_pattern.search(course_text)
                if instructor_match:
                    instructor = instructor_match.group(1)
                
                if course_name not in current_courses:
                    current_courses[course_name] = {
                        'title': course_name,
                        'description': '',
                        'instructor': instructor,
                        'classroom': classroom,
                        'schedule': []
                    }
                
                schedule_entry = {
                    'dayOfWeek': current_day,
                    'startTime': start_time,
                    'endTime': end_time
                }
                current_courses[course_name]['schedule'].append(schedule_entry)
    
    # Convert to list
    for course_data in current_courses.values():
        if course_data['schedule']:
            courses.append({
                'title': course_data['title'][:255],
                'description': course_data['description'][:1000],
                'instructor': course_data['instructor'][:100],
                'classroom': course_data['classroom'][:100],
                'schedule': course_data['schedule']
            })
    
    return courses[:15]


